import pandas as pd
import os
import glob
import json
import tempfile
import shutil
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# S3 Configuration
USE_S3 = os.getenv('USE_S3', 'false').lower() == 'true'
S3_BUCKET_NAME = os.getenv('S3_BUCKET_NAME')
S3_BASE_PREFIX = os.getenv('S3_BASE_PREFIX', 'trade_surveillance')

# Check if S3 is available (boto3 imported successfully)
S3_AVAILABLE = False
try:
    import boto3
    from s3_utils import (
        s3_file_exists, read_excel_from_s3, read_csv_from_s3,
        read_text_from_s3, read_json_from_s3, upload_file_to_s3,
        download_file_from_s3, list_s3_objects
    )
    S3_AVAILABLE = True
except ImportError:
    print("Boto3 not installed, S3 functions will not be available.")
except Exception as e:
    print(f"Error importing S3 utilities: {e}")

def add_required_columns_for_date(date_str):
    """
    Add required columns to the final analysis file for a specific date in August or September
    date_str format: '01082025' for August 1st, 2025 or '01092025' for September 1st, 2025
    """
    
    # Determine month and set paths accordingly
    month = int(date_str[2:4])
    month_names = {
        1: "January", 2: "February", 3: "March", 4: "April",
        5: "May", 6: "June", 7: "July", 8: "August",
        9: "September", 10: "October", 11: "November", 12: "December"
    }
    
    if month not in month_names:
        print(f"❌ Invalid month: {month}")
        return None
    
    month_name = month_names[month]
    
    # Define paths based on month
    analysis_file = f"{month_name}/Daily_Reports/{date_str}/order_transcript_analysis_{date_str}.xlsx"
    order_files_path = f"{month_name}/Order Files"
    call_info_file = f"{month_name}/Daily_Reports/{date_str}/call_info_output_{date_str}.xlsx"
    audio_validation_file = f"{month_name}/Daily_Reports/{date_str}/audio_order_kl_orgtimestamp_validation_{date_str}.xlsx"
    transcripts_path = f"{month_name}/Daily_Reports/{date_str}/transcripts_{date_str}"
    email_mapping_file = f"{month_name}/Daily_Reports/{date_str}/email_order_mapping_{date_str}.json"
    output_path = f"{month_name}/Daily_Reports/{date_str}/order_transcript_analysis_{date_str}_with_required_columns.xlsx"
    
    # Get S3 keys if using S3
    if USE_S3 and S3_AVAILABLE:
        analysis_s3_key = f"{S3_BASE_PREFIX}/{analysis_file}"
        call_info_s3_key = f"{S3_BASE_PREFIX}/{call_info_file}"
        audio_validation_s3_key = f"{S3_BASE_PREFIX}/{audio_validation_file}"
        transcripts_s3_prefix = f"{S3_BASE_PREFIX}/{transcripts_path}/"
        email_mapping_s3_key = f"{S3_BASE_PREFIX}/{email_mapping_file}"
        output_s3_key = f"{S3_BASE_PREFIX}/{output_path}"
        
        # Check if analysis file exists
        if not s3_file_exists(analysis_s3_key):
            print(f"Analysis file not found in S3: {analysis_s3_key}")
            print("Note: This step requires Step 6 (AI Analysis) to be completed first")
            return None
        
        # Load the analysis file from S3
        try:
            df = read_excel_from_s3(analysis_s3_key)
            print(f"Loaded analysis file from S3 with {len(df)} records")
        except Exception as e:
            print(f"Error loading analysis file from S3: {e}")
            import traceback
            traceback.print_exc()
            return None
        
        # Load call info from S3
        try:
            call_info_df = read_excel_from_s3(call_info_s3_key)
            print(f"Loaded call info from S3 with {len(call_info_df)} records")
        except Exception as e:
            print(f"Error loading call info from S3: {e}")
            import traceback
            traceback.print_exc()
            return None
        
        # Find order file for the specific date in S3
        order_file_patterns = [
            f"{S3_BASE_PREFIX}/{order_files_path}/OrderBook-Closed-{date_str}.csv",
            f"{S3_BASE_PREFIX}/{month_name}/Daily_Reports/{date_str}/OrderBook-Closed-{date_str}.csv",
            f"{S3_BASE_PREFIX}/{order_files_path}/OrderBook_Closed-{date_str}.csv",
            f"{S3_BASE_PREFIX}/{month_name}/Daily_Reports/{date_str}/OrderBook_Closed-{date_str}.csv"
        ]
        
        order_file_s3_key = None
        for pattern_key in order_file_patterns:
            if s3_file_exists(pattern_key):
                order_file_s3_key = pattern_key
                print(f"Found order file in S3: {order_file_s3_key}")
                break
        
        if not order_file_s3_key:
            print(f"Order file not found in S3 for any pattern: {order_file_patterns}")
            return None
        
        # Load order data from S3
        try:
            order_df = read_csv_from_s3(order_file_s3_key)
            print(f"Loaded order file from S3 with {len(order_df)} records")
        except Exception as e:
            print(f"Error loading order file from S3: {e}")
            import traceback
            traceback.print_exc()
            return None
        
        # Load audio validation data if available from S3
        audio_validation_df = None
        if s3_file_exists(audio_validation_s3_key):
            try:
                # Download and read specific sheet
                temp_file_path = download_file_from_s3(audio_validation_s3_key)
                audio_validation_df = pd.read_excel(temp_file_path, sheet_name='Order_Audio_Mapping')
                os.unlink(temp_file_path)
                print(f"Loaded audio validation data from S3 with {len(audio_validation_df)} records")
                print(f"Audio matches found: {len(audio_validation_df[audio_validation_df['has_audio'] == 'Y'])}")
            except Exception as e:
                print(f"Error loading audio validation data from S3: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("Audio validation file not found in S3 - proceeding without audio validation data")
        
        # Load email mapping data if available from S3
        email_mapping_data = None
        if s3_file_exists(email_mapping_s3_key):
            try:
                email_mapping_data = read_json_from_s3(email_mapping_s3_key)
                print(f"Loaded email mapping data from S3 with {len(email_mapping_data.get('matches', []))} matches")
            except Exception as e:
                print(f"Error loading email mapping data from S3: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("Email mapping file not found in S3 - proceeding without email data")
    else:
        # Local filesystem
        # Check if analysis file exists
        if not os.path.exists(analysis_file):
            print(f"Analysis file not found: {analysis_file}")
            print("Note: This step requires Step 6 (AI Analysis) to be completed first")
            return None
        
        # Load the analysis file
        try:
            df = pd.read_excel(analysis_file)
            print(f"Loaded analysis file with {len(df)} records")
        except Exception as e:
            print(f"Error loading analysis file: {e}")
            return None
        
        # Load call info
        try:
            call_info_df = pd.read_excel(call_info_file)
            print(f"Loaded call info with {len(call_info_df)} records")
        except Exception as e:
            print(f"Error loading call info: {e}")
            return None
        
        # Find order file for the specific date
        order_file_pattern = f"OrderBook-Closed-{date_str}.csv"
        order_file_path = os.path.join(order_files_path, order_file_pattern)
        
        if not os.path.exists(order_file_path):
            print(f"Order file not found: {order_file_path}")
            return None
        
        # Load order data
        try:
            order_df = pd.read_csv(order_file_path)
            print(f"Loaded order file with {len(order_df)} records")
        except Exception as e:
            print(f"Error loading order file: {e}")
            return None
        
        # Load audio validation data if available
        audio_validation_df = None
        if os.path.exists(audio_validation_file):
            try:
                audio_validation_df = pd.read_excel(audio_validation_file, sheet_name='Order_Audio_Mapping')
                print(f"Loaded audio validation data with {len(audio_validation_df)} records")
                print(f"Audio matches found: {len(audio_validation_df[audio_validation_df['has_audio'] == 'Y'])}")
            except Exception as e:
                print(f"Error loading audio validation data: {e}")
        else:
            print("Audio validation file not found - proceeding without audio validation data")
        
        # Load email mapping data if available
        email_mapping_data = None
        if os.path.exists(email_mapping_file):
            try:
                with open(email_mapping_file, 'r') as f:
                    email_mapping_data = json.load(f)
                print(f"Loaded email mapping data with {len(email_mapping_data.get('matches', []))} matches")
            except Exception as e:
                print(f"Error loading email mapping data: {e}")
        else:
            print("Email mapping file not found - proceeding without email data")
    
    # Add required columns
    print("Adding required columns...")
    
    # 1. Order Date - extract from order_time or use current date
    df['Order Date'] = pd.to_datetime(df['order_time']).dt.date if 'order_time' in df.columns else pd.Timestamp.now().date()
    
    # 2. Order ID - already present
    df['Order ID'] = df['order_id']
    
    # 3. Client Code - already present
    df['Client Code'] = df['client_id']
    
    # 4. Dealer ID - from order files
    df['Dealer ID'] = df['user']
    
    # 5. Mobile No. - from call info (map via audio_file)
    audio_to_mobile = dict(zip(call_info_df['filename'], call_info_df['mobile_number']))
    df['Mobile No.'] = df['audio_file'].map(audio_to_mobile)
    
    # 6. Call received from Registered Number (Y/N) - from call info
    audio_to_registered = dict(zip(call_info_df['filename'], call_info_df['present_in_ucc']))
    df['Call received from Registered Number (Y/N)'] = df['audio_file'].map(audio_to_registered)
    
    # 7. Call Records matched with Order File (Y/N) - from audio validation
    if audio_validation_df is not None:
        # PERMANENT FIX: Use EXACT STRING MATCHING after normalization
        # Excel converts large integers to floats/scientific notation, but we can normalize to exact strings
        # This is the same approach used in OMS validation (proven to work)
        
        def normalize_order_id_to_string(val):
            """Normalize order ID to exact string for matching (handles scientific notation)"""
            if pd.isna(val):
                return None
            try:
                # Convert scientific notation/float to full integer, then to string
                # This gives us exact matching regardless of Excel's float conversion
                return str(int(float(val)))
            except (ValueError, TypeError):
                # Fallback: remove trailing .0 if present
                s = str(val)
                return s[:-2] if s.endswith('.0') else s
        
        # Build mapping dictionaries using normalized string keys
        order_id_to_audio = {}
        order_id_to_files = {}
        
        for _, row in audio_validation_df.iterrows():
            order_id = row.get('order_id')
            has_audio = row.get('has_audio', 'N')
            mapped_files = str(row.get('mapped_audio_filenames', ''))
            
            # Skip NaN order IDs
            if pd.isna(order_id):
                continue
            
            # Normalize to string for exact matching
            order_id_str = normalize_order_id_to_string(order_id)
            if order_id_str is None:
                continue
            
            if has_audio == 'Y':
                order_id_to_audio[order_id_str] = 'yes'
                if mapped_files and mapped_files != 'nan' and mapped_files.strip():
                    order_id_to_files[order_id_str] = mapped_files
        
        # Match using exact string comparison (normalized)
        def get_audio_match_and_file(order_id_val):
            """Get both audio match status and file name using EXACT string matching"""
            if pd.isna(order_id_val):
                return ('no', None)
            
            # Normalize to string for exact matching
            order_id_str = normalize_order_id_to_string(order_id_val)
            if order_id_str is None:
                return ('no', None)
            
            # Exact string match - simple and reliable
            if order_id_str in order_id_to_audio and order_id_str in order_id_to_files:
                return (order_id_to_audio[order_id_str], order_id_to_files[order_id_str])
            
            return ('no', None)
        
        # Apply matching and get both status and file
        audio_results = df['order_id'].apply(get_audio_match_and_file)
        df['Call Records matched with Order File (Y/N)'] = audio_results.apply(lambda x: x[0])
        df['audio_mapped'] = df['Call Records matched with Order File (Y/N)'].apply(lambda x: 'yes' if x == 'yes' else 'no')
        
        # Update audio_file column with mapped filenames - ONLY if audio_mapped='yes'
        # PRESERVE EXTENSION: Keep full filename with .mp3/.wav extension for proper file matching
        for idx, (audio_status, audio_file) in enumerate(audio_results):
            if audio_status == 'yes' and audio_file:
                # Take first file if multiple files (comma-separated), preserve full filename with extension
                audio_file_value = audio_file.split(',')[0].strip()
                df.at[df.index[idx], 'audio_file'] = audio_file_value
        
        # PERMANENT FIX: Handle orders with NaN order_id that have audio files
        # These orders were matched in audio validation but can't be matched by order_id
        # If an order has a Call File Name (from intermediate file) but audio_mapped='no',
        # check if that audio file exists in audio validation and set audio_mapped='yes'
        
        # Build set of all audio files that have matches in validation
        all_matched_audio_files = set()
        for order_id_str, audio_file in order_id_to_files.items():
            if audio_file:
                # Handle comma-separated files
                for f in audio_file.split(','):
                    all_matched_audio_files.add(f.strip())
        
        # Also check for orders with NaN order_id in validation that have audio files
        nan_orders_with_audio = audio_validation_df[
            (pd.isna(audio_validation_df['order_id'])) & 
            (audio_validation_df['has_audio'] == 'Y') &
            (audio_validation_df['mapped_audio_filenames'].notna())
        ]
        for _, row in nan_orders_with_audio.iterrows():
            audio_files = str(row.get('mapped_audio_filenames', ''))
            if audio_files and audio_files != 'nan':
                for f in audio_files.split(','):
                    all_matched_audio_files.add(f.strip())
        
        # PERMANENT FIX: Handle orders that have audio files but weren't matched by order_id
        # This includes:
        # 1. Orders with NaN order_id (can't be matched by ID)
        # 2. Orders with valid order_id but not in audio validation (matched via transcript analysis in Step 6)
        # If an order has a Call File Name, check if that audio file exists in call info
        # If it does, mark as matched (the audio file exists and is associated with the order)
        
        # Build set of all valid audio files from call info
        valid_audio_files = set()
        if USE_S3 and S3_AVAILABLE:
            if s3_file_exists(call_info_s3_key):
                try:
                    call_info_df_check = read_excel_from_s3(call_info_s3_key)
                    if 'filename' in call_info_df_check.columns:
                        valid_audio_files = set(call_info_df_check['filename'].dropna().astype(str))
                except Exception as e:
                    print(f"Warning: Could not read call info from S3 to validate audio files: {e}")
        else:
            if os.path.exists(call_info_file):
                try:
                    call_info_df_check = pd.read_excel(call_info_file)
                    if 'filename' in call_info_df_check.columns:
                        valid_audio_files = set(call_info_df_check['filename'].dropna().astype(str))
                except Exception as e:
                    print(f"Warning: Could not read call info to validate audio files: {e}")
        
        # For orders with audio file but audio_mapped='no', check if file is valid
        for idx, row in df.iterrows():
            if row.get('audio_mapped') == 'no':
                # Check if this order has an audio file (from intermediate file or already set)
                existing_audio_file = row.get('audio_file') or row.get('Call File Name', '')
                if existing_audio_file and str(existing_audio_file) != 'nan':
                    audio_file_clean = str(existing_audio_file).strip()
                    
                    # Check if this audio file is in the matched set (from validation)
                    if audio_file_clean in all_matched_audio_files:
                        # This audio file was matched in validation, so mark as matched
                        df.at[idx, 'audio_mapped'] = 'yes'
                        df.at[idx, 'Call Records matched with Order File (Y/N)'] = 'yes'
                        if pd.isna(row.get('order_id')):
                            print(f"Fixed NaN order_id match: audio file {audio_file_clean} was matched in validation")
                    
                    # Also check if audio file exists in call info (from Step 6 transcript analysis)
                    elif audio_file_clean in valid_audio_files:
                        # Audio file exists in call info, so it's a valid association
                        # Mark as matched (order was associated via transcript analysis)
                        df.at[idx, 'audio_mapped'] = 'yes'
                        df.at[idx, 'Call Records matched with Order File (Y/N)'] = 'yes'
                        print(f"Fixed audio match: audio file {audio_file_clean} exists in call info (from transcript analysis)")
    else:
        # Fallback to order_discussed if audio validation not available
        df['Call Records matched with Order File (Y/N)'] = df.get('order_discussed', 'no')
    
    # 8. Order Executed (Y/N) - from order files
    # Create a proper mapping from order file
    order_status_map = {}
    for _, row in order_df.iterrows():
        order_id = row['ExchOrderID']
        status = row['Status']
        if order_id not in order_status_map:  # Take first occurrence if duplicates
            order_status_map[order_id] = status
    
    df['Order Executed (Y/N)'] = df['order_id'].map(lambda x: 'Y' if order_status_map.get(x) == 'Complete' else 'N')
    
    # 9. Call Extract - from transcripts
    df['Call Extract'] = ''
    for idx, row in df.iterrows():
        audio_file = row['audio_file']
        # Skip if audio_file is NaN or empty
        if pd.isna(audio_file) or audio_file == '':
            continue
        
        # Handle composite audio filenames (multiple files joined with commas)
        if ',' in audio_file:
            # Split composite filename and read all individual transcript files
            individual_files = [f.strip() for f in audio_file.split(',')]
            combined_transcript = ""
            for individual_file in individual_files:
                if USE_S3 and S3_AVAILABLE:
                    transcript_s3_key = f"{transcripts_s3_prefix}{individual_file}.txt"
                    if s3_file_exists(transcript_s3_key):
                        try:
                            transcript = read_text_from_s3(transcript_s3_key)
                            if combined_transcript:
                                combined_transcript += f"\n\n=== {individual_file} ===\n"
                            combined_transcript += transcript
                        except Exception as e:
                            print(f"Error reading transcript from S3 for {individual_file}: {e}")
                else:
                    transcript_file = os.path.join(transcripts_path, individual_file + ".txt")
                    if os.path.exists(transcript_file):
                        try:
                            with open(transcript_file, 'r', encoding='utf-8') as f:
                                transcript = f.read()
                            if combined_transcript:
                                combined_transcript += f"\n\n=== {individual_file} ===\n"
                            combined_transcript += transcript
                        except Exception as e:
                            print(f"Error reading transcript for {individual_file}: {e}")
            
            if combined_transcript:
                df.at[idx, 'Call Extract'] = combined_transcript[:1000] + "..." if len(combined_transcript) > 1000 else combined_transcript
        else:
            # Single audio file
            if USE_S3 and S3_AVAILABLE:
                transcript_s3_key = f"{transcripts_s3_prefix}{audio_file}.txt"
                if s3_file_exists(transcript_s3_key):
                    try:
                        transcript = read_text_from_s3(transcript_s3_key)
                        df.at[idx, 'Call Extract'] = transcript[:1000] + "..." if len(transcript) > 1000 else transcript
                    except Exception as e:
                        print(f"Error reading transcript from S3 for {audio_file}: {e}")
            else:
                transcript_file = os.path.join(transcripts_path, audio_file + ".txt")
                if os.path.exists(transcript_file):
                    try:
                        with open(transcript_file, 'r', encoding='utf-8') as f:
                            transcript = f.read()
                        df.at[idx, 'Call Extract'] = transcript[:1000] + "..." if len(transcript) > 1000 else transcript
                    except Exception as e:
                        print(f"Error reading transcript for {audio_file}: {e}")
    
    # 10. Call File Name - already present
    df['Call File Name'] = df['audio_file']
    
    # 11. Observation - from AI analysis
    df['Observation'] = df['ai_reasoning']
    
    # 12. Email-Order Match Status - from email mapping (only for actually matched orders)
    # PERMANENT FIX: Apply OMS matches from intermediate file (if Step 4 ran before Step 9)
    # OR preserve OMS_MATCH status from existing Excel (if Step 4 ran after Step 9)
    
    # PERMANENT FIX: Check for OMS matches intermediate file first
    # This handles the case where Step 8 ran before Step 9 (Excel didn't exist yet)
    # FIX: Use absolute path to ensure file is found regardless of current working directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    oms_matches_file = os.path.join(base_dir, 'oms_surveillance', f'oms_matches_{date_str}.json')
    existing_oms_matches = {}
    
    # Check for OMS matches file (local or S3)
    if USE_S3 and S3_AVAILABLE:
        # OMS matches file is stored locally in the container, not in S3
        # So we still check the local path
        oms_matches_file_local = oms_matches_file
    else:
        oms_matches_file_local = oms_matches_file
    
    print(f"🔍 [STEP9] Looking for OMS matches file: {oms_matches_file_local}")
    print(f"🔍 [STEP9] File exists: {os.path.exists(oms_matches_file_local)}")
    print(f"🔍 [STEP9] Current working directory: {os.getcwd()}")
    print(f"🔍 [STEP9] Base directory: {base_dir}")
    
    if os.path.exists(oms_matches_file_local):
        try:
            print(f"📋 Found OMS matches intermediate file: {oms_matches_file_local}")
            with open(oms_matches_file_local, 'r', encoding='utf-8') as f:
                oms_matches_data = json.load(f)
            
            matches = oms_matches_data.get('matches', {})
            if matches:
                print(f"✅ Loading {len(matches)} OMS matches from intermediate file")
                
                # Normalize order IDs for matching
                def normalize_order_id_to_string(val):
                    if pd.isna(val):
                        return None
                    try:
                        return str(int(float(val)))
                    except (ValueError, TypeError):
                        s = str(val)
                        return s[:-2] if s.endswith('.0') else s
                
                # Create mapping of order ID to OMS_MATCH status
                for order_id, match_data in matches.items():
                    normalized_id = normalize_order_id_to_string(order_id)
                    if normalized_id:
                        existing_oms_matches[normalized_id] = 'OMS_MATCH'
                
                print(f"✅ Prepared {len(existing_oms_matches)} OMS matches to apply")
            else:
                print(f"⚠️ [STEP9] Intermediate file exists but contains no matches")
        except Exception as e:
            print(f"⚠️  Warning: Could not load OMS matches from intermediate file: {e}")
            import traceback
            print(f"⚠️ [STEP9] Traceback: {traceback.format_exc()}")
    else:
        print(f"⚠️ [STEP9] OMS matches intermediate file NOT found: {oms_matches_file_local}")
        print(f"⚠️ [STEP9] This means either:")
        print(f"   1. Step 8 didn't run yet")
        print(f"   2. Step 8 ran but Excel existed, so intermediate file was deleted")
        print(f"   3. Step 8 ran but file path is incorrect")
    
    # PERMANENT FIX: Always check if final Excel file already exists (from Step 8 OMS update)
    # This handles the case where Step 8 runs AFTER Step 9 (intermediate file was deleted)
    # OR where Step 8 runs BEFORE Step 9 (intermediate file might not exist)
    final_report_name = f"Final_Trade_Surveillance_Report_{date_str}_with_Email_and_Trade_Analysis.xlsx"
    final_report_path = os.path.join(os.path.dirname(output_path), final_report_name)
    
    # Get S3 key for final report if using S3
    if USE_S3 and S3_AVAILABLE:
        final_report_s3_key = f"{S3_BASE_PREFIX}/{os.path.dirname(output_path)}/{final_report_name}"
    else:
        final_report_s3_key = None
    
    # FIX: Always check existing Excel file, not just when existing_oms_matches is empty
    # This ensures OMS matches are preserved even if Step 8 ran after Step 9
    if USE_S3 and S3_AVAILABLE:
        if s3_file_exists(final_report_s3_key):
            try:
                df_existing = read_excel_from_s3(final_report_s3_key)
                if 'Order ID' in df_existing.columns and 'Email-Order Match Status' in df_existing.columns:
                    # Create mapping of order ID to OMS_MATCH status
                    oms_matched = df_existing[df_existing['Email-Order Match Status'] == 'OMS_MATCH']
                    if len(oms_matched) > 0:
                        print(f"📋 Found {len(oms_matched)} existing OMS matches in Final Excel file from S3")
                        # Normalize order IDs for matching
                        def normalize_order_id_to_string(val):
                            if pd.isna(val):
                                return None
                            try:
                                return str(int(float(val)))
                            except (ValueError, TypeError):
                                s = str(val)
                                return s[:-2] if s.endswith('.0') else s
                        
                        # Merge with existing_oms_matches (don't overwrite if already found from intermediate file)
                        for idx, row in oms_matched.iterrows():
                            order_id = normalize_order_id_to_string(row['Order ID'])
                            if order_id and order_id not in existing_oms_matches:
                                existing_oms_matches[order_id] = 'OMS_MATCH'
                        
                        if len(existing_oms_matches) > 0:
                            print(f"✅ Preserving {len(existing_oms_matches)} OMS matches from Step 8 (from Final Excel in S3)")
            except Exception as e:
                print(f"⚠️  Warning: Could not read existing Excel file from S3 to preserve OMS matches: {e}")
    else:
        if os.path.exists(final_report_path):
            try:
                df_existing = pd.read_excel(final_report_path)
                if 'Order ID' in df_existing.columns and 'Email-Order Match Status' in df_existing.columns:
                    # Create mapping of order ID to OMS_MATCH status
                    oms_matched = df_existing[df_existing['Email-Order Match Status'] == 'OMS_MATCH']
                    if len(oms_matched) > 0:
                        print(f"📋 Found {len(oms_matched)} existing OMS matches in Final Excel file")
                        # Normalize order IDs for matching
                        def normalize_order_id_to_string(val):
                            if pd.isna(val):
                                return None
                            try:
                                return str(int(float(val)))
                            except (ValueError, TypeError):
                                s = str(val)
                                return s[:-2] if s.endswith('.0') else s
                        
                        # Merge with existing_oms_matches (don't overwrite if already found from intermediate file)
                        for idx, row in oms_matched.iterrows():
                            order_id = normalize_order_id_to_string(row['Order ID'])
                            if order_id and order_id not in existing_oms_matches:
                                existing_oms_matches[order_id] = 'OMS_MATCH'
                        
                        if len(existing_oms_matches) > 0:
                            print(f"✅ Preserving {len(existing_oms_matches)} OMS matches from Step 8 (from Final Excel)")
            except Exception as e:
                print(f"⚠️  Warning: Could not read existing Excel file to preserve OMS matches: {e}")
    
    # PERMANENT FIX: Preserve existing OMS_MATCH status from source file
    # If the analysis file already has OMS_MATCH status (from Step 8), preserve it
    if 'Email-Order Match Status' in df.columns:
        existing_oms_status = df[df['Email-Order Match Status'] == 'OMS_MATCH'].copy()
        if len(existing_oms_status) > 0:
            print(f"📋 Found {len(existing_oms_status)} existing OMS_MATCH statuses in source file - will preserve")
            # Store these for later restoration
            existing_oms_order_ids = set()
            def normalize_order_id_for_preservation(val):
                if pd.isna(val):
                    return None
                try:
                    return str(int(float(val)))
                except (ValueError, TypeError):
                    s = str(val)
                    return s[:-2] if s.endswith('.0') else s
            existing_oms_status['Order ID Normalized'] = existing_oms_status['Order ID'].apply(normalize_order_id_for_preservation)
            existing_oms_order_ids = set(existing_oms_status['Order ID Normalized'].dropna().tolist())
    else:
        existing_oms_order_ids = set()
    
    # Initialize with 'No Email Match' default
    df['Email-Order Match Status'] = 'No Email Match'
    
    # First, apply OMS matches (from intermediate file or existing Excel)
    if existing_oms_matches:
        print(f"✅ Applying {len(existing_oms_matches)} OMS matches to Excel")
        def normalize_order_id_for_matching(val):
            if pd.isna(val):
                return None
            try:
                return str(int(float(val)))
            except (ValueError, TypeError):
                s = str(val)
                return s[:-2] if s.endswith('.0') else s
        
        df['Order ID Normalized'] = df['Order ID'].apply(normalize_order_id_for_matching)
        df['Email-Order Match Status'] = df['Order ID Normalized'].map(existing_oms_matches).fillna(df['Email-Order Match Status'])
        df = df.drop('Order ID Normalized', axis=1)  # Remove temporary column
    
    # PERMANENT FIX: Restore OMS_MATCH status from source file if it existed
    if existing_oms_order_ids:
        print(f"✅ Restoring {len(existing_oms_order_ids)} OMS_MATCH statuses from source file")
        def normalize_order_id_for_restore(val):
            if pd.isna(val):
                return None
            try:
                return str(int(float(val)))
            except (ValueError, TypeError):
                s = str(val)
                return s[:-2] if s.endswith('.0') else s
        df['Order ID Normalized'] = df['Order ID'].apply(normalize_order_id_for_restore)
        df.loc[df['Order ID Normalized'].isin(existing_oms_order_ids), 'Email-Order Match Status'] = 'OMS_MATCH'
        df = df.drop('Order ID Normalized', axis=1)  # Remove temporary column
    
    # Then, add email matches (but don't overwrite OMS_MATCH)
    if email_mapping_data:
        # Create mapping from order ID to email match status (only for actually matched orders)
        order_id_to_email_status = {}
        for match in email_mapping_data.get('matches', []):
            email_instruction = match.get('email_instruction', {})
            matched_orders = match.get('matched_orders', [])
            
            # Only process if there are actually matched orders
            if matched_orders:
                # Determine match status based on match type and confidence
                match_type = match.get('match_type', 'NO_MATCH')
                confidence_score = match.get('confidence_score', 0)
                
                # CRITICAL FIX: ORDER_CONFLICT should still be marked as 'Matched'
                # The conflict flag is just for review, not a reason to mark as unmatched
                # If matched_orders is not empty, it means orders were matched (regardless of conflict)
                if match_type in ['EXACT_MATCH', 'HIGH_CONFIDENCE_MATCH'] and confidence_score >= 80:
                    match_status = 'Matched'
                elif match_type == 'SPLIT_EXECUTION':
                    match_status = 'Matched'
                elif match_type == 'ORDER_CONFLICT':
                    # ORDER_CONFLICT means orders were matched but assigned to multiple instructions
                    # Still mark as 'Matched' since orders were successfully matched
                    match_status = 'Matched'
                elif match_type == 'PARTIAL_MATCH' and confidence_score >= 70:
                    match_status = 'Partial Match'
                else:
                    # For any other match type with matched_orders, still mark as matched if confidence is high
                    if confidence_score >= 80:
                        match_status = 'Matched'
                    else:
                        match_status = 'No Match'
                
                # Map to each matched order
                for matched_order in matched_orders:
                    order_id = matched_order.get('ExchOrderID')
                    if order_id:
                        order_id_to_email_status[order_id] = match_status
        
        # Map email matches, but preserve OMS_MATCH (don't overwrite)
        def apply_email_match_status(row):
            current_status = row['Email-Order Match Status']
            # If already OMS_MATCH, preserve it
            if current_status == 'OMS_MATCH':
                return 'OMS_MATCH'
            # Otherwise, apply email match status
            order_id = row.get('Order ID')
            if order_id and order_id in order_id_to_email_status:
                return order_id_to_email_status[order_id]
            return current_status
        
        df['Email-Order Match Status'] = df.apply(apply_email_match_status, axis=1)
    else:
        # If no email mapping data, keep existing status (including OMS_MATCH if preserved)
        pass
    
    # 13. Email Confidence Score - from email mapping (only for actually matched orders)
    df['Email Confidence Score'] = '0%'  # Default value
    if email_mapping_data:
        # Create mapping from order ID to email confidence score (only for actually matched orders)
        order_id_to_email_confidence = {}
        for match in email_mapping_data.get('matches', []):
            email_instruction = match.get('email_instruction', {})
            matched_orders = match.get('matched_orders', [])
            
            # Only process if there are actually matched orders
            if matched_orders:
                confidence_score = match.get('confidence_score', 0)
                # Format as percentage if it's a number
                if isinstance(confidence_score, (int, float)):
                    confidence_str = f"{confidence_score}%"
                else:
                    confidence_str = f"{confidence_score}%"
                
                # Map to each matched order
                for matched_order in matched_orders:
                    order_id = matched_order.get('ExchOrderID')
                    if order_id:
                        order_id_to_email_confidence[order_id] = confidence_str
        
        # Map to orders based on order ID (only for actually matched orders)
        df['Email Confidence Score'] = df['Order ID'].map(order_id_to_email_confidence).fillna('0%')
    else:
        df['Email Confidence Score'] = '0%'
    
    # 14. Email Discrepancy Details - from email mapping (only for actually matched orders)
    df['Email Discrepancy Details'] = ''
    if email_mapping_data:
        # Create mapping from order ID to email discrepancy details (only for actually matched orders)
        order_id_to_email_discrepancies = {}
        for match in email_mapping_data.get('matches', []):
            email_instruction = match.get('email_instruction', {})
            matched_orders = match.get('matched_orders', [])
            
            # Only process if there are actually matched orders
            if matched_orders:
                discrepancies = match.get('discrepancies', [])
                if discrepancies:
                    discrepancy_str = '; '.join(discrepancies)
                else:
                    discrepancy_str = 'No discrepancies'
                
                # Map to each matched order
                for matched_order in matched_orders:
                    order_id = matched_order.get('ExchOrderID')
                    if order_id:
                        order_id_to_email_discrepancies[order_id] = discrepancy_str
        
        # Map to orders based on order ID (only for actually matched orders)
        df['Email Discrepancy Details'] = df['Order ID'].map(order_id_to_email_discrepancies).fillna('No email data')
    else:
        df['Email Discrepancy Details'] = 'No email data'
    
    # 15. Email Content - from email mapping (only for actually matched orders)
    df['Email_Content'] = ''
    if email_mapping_data:
        # Create mapping from order ID to email content (only for matched orders)
        order_id_to_email_content = {}
        for match in email_mapping_data.get('matches', []):
            email_instruction = match.get('email_instruction', {})
            matched_orders = match.get('matched_orders', [])
            
            # Only process if there are actually matched orders
            if matched_orders:
                # Get the full email content
                source_email = email_instruction.get('source_email', {})
                full_email_content = source_email.get('clean_text', '')
                
                # Map to each matched order
                for matched_order in matched_orders:
                    order_id = matched_order.get('ExchOrderID')
                    if order_id:
                        order_id_to_email_content[order_id] = full_email_content
        
        # Map to orders based on order ID (only for actually matched orders)
        df['Email_Content'] = df['Order ID'].map(order_id_to_email_content).fillna('')
    else:
        df['Email_Content'] = ''
    
    # Reorder columns to match required format
    required_columns = [
        'Order Date', 'Order ID', 'Client Code', 'Dealer ID', 'Mobile No.',
        'Call received from Registered Number (Y/N)', 'Call Records matched with Order File (Y/N)',
        'Order Executed (Y/N)', 'Call Extract', 'Call File Name', 'Observation',
        'Email-Order Match Status', 'Email Confidence Score', 'Email Discrepancy Details', 'Email_Content'
    ]
    
    # Add any missing required columns with empty values
    for col in required_columns:
        if col not in df.columns:
            df[col] = ''
    
    # Reorder columns and add original analysis columns
    final_columns = required_columns + [col for col in df.columns if col not in required_columns]
    df_final = df[final_columns]
    
    # Add highlighting for orders with no source (no audio + no email + completed status)
    print("Adding highlighting for orders with no source...")
    
    # Import openpyxl for highlighting
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        # Save the final file with enhanced naming
        final_report_name = f"Final_Trade_Surveillance_Report_{date_str}_with_Email_and_Trade_Analysis.xlsx"
        final_report_path = os.path.join(os.path.dirname(output_path), final_report_name)
        
        # Save to temporary file first (for highlighting)
        temp_output_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        df_final.to_excel(temp_output_file.name, index=False)
        
        # Load the workbook for highlighting
        workbook = load_workbook(temp_output_file.name)
        worksheet = workbook.active
        
        # Define highlighting styles
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        
        # Apply highlighting to rows with no source
        no_source_count = 0
        for row_idx, row in df_final.iterrows():
            # Check if order has no audio, no email, and is completed
            no_audio = row['Call Records matched with Order File (Y/N)'] != 'yes'
            no_email = row['Email-Order Match Status'] != 'Matched'
            completed = row['Order Executed (Y/N)'] == 'Y'
            
            if no_audio and no_email and completed:
                # Highlight the entire row (Excel rows are 1-indexed, and we have header)
                excel_row = row_idx + 2  # +2 because Excel is 1-indexed and we have header
                for col in range(1, len(df_final.columns) + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.fill = red_fill
                no_source_count += 1
        
        # Save the updated file with highlighting
        workbook.save(temp_output_file.name)
        
        # Upload to S3 or save locally
        if USE_S3 and S3_AVAILABLE:
            upload_file_to_s3(temp_output_file.name, final_report_s3_key)
            print(f"Final comprehensive trade surveillance report saved to S3: {final_report_s3_key}")
        else:
            shutil.move(temp_output_file.name, final_report_path)
            print(f"Final comprehensive trade surveillance report saved to: {final_report_path}")
        
        # Clean up temp file
        if os.path.exists(temp_output_file.name):
            os.unlink(temp_output_file.name)
        
        # PERMANENT FIX: Delete OMS matches intermediate file if it exists and matches were applied
        if os.path.exists(oms_matches_file_local) and existing_oms_matches:
            try:
                os.remove(oms_matches_file_local)
                print(f"🗑️  Deleted OMS matches intermediate file (matches applied to Excel)")
            except Exception as e:
                print(f"⚠️  Warning: Could not delete OMS matches intermediate file: {e}")
        
        print(f"Total records: {len(df_final)}")
        print(f"⚠️  Orders with no source (highlighted in red): {no_source_count}")
        
    except ImportError:
        print("openpyxl not available - saving without highlighting")
        final_report_name = f"Final_Trade_Surveillance_Report_{date_str}_with_Email_and_Trade_Analysis.xlsx"
        final_report_path = os.path.join(os.path.dirname(output_path), final_report_name)
        
        # Save to temporary file first
        temp_output_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        df_final.to_excel(temp_output_file.name, index=False)
        
        # Upload to S3 or save locally
        if USE_S3 and S3_AVAILABLE:
            upload_file_to_s3(temp_output_file.name, final_report_s3_key)
            print(f"Final comprehensive trade surveillance report saved to S3: {final_report_s3_key}")
        else:
            shutil.move(temp_output_file.name, final_report_path)
            print(f"Final comprehensive trade surveillance report saved to: {final_report_path}")
        
        # Clean up temp file
        if os.path.exists(temp_output_file.name):
            os.unlink(temp_output_file.name)
        
        # PERMANENT FIX: Delete OMS matches intermediate file if it exists and matches were applied
        if os.path.exists(oms_matches_file_local) and existing_oms_matches:
            try:
                os.remove(oms_matches_file_local)
                print(f"🗑️  Deleted OMS matches intermediate file (matches applied to Excel)")
            except Exception as e:
                print(f"⚠️  Warning: Could not delete OMS matches intermediate file: {e}")
        
        print(f"Total records: {len(df_final)}")
    
    # Print summary with correct logic
    print("\nSummary:")
    print(f"- Orders with audio: {len(df_final)}")
    print(f"- Orders discussed: {len(df_final[df_final['Call Records matched with Order File (Y/N)'] == 'yes'])}")
    print(f"- Orders executed: {len(df_final[df_final['Order Executed (Y/N)'] == 'Y'])}")
    print(f"- Registered numbers: {len(df_final[df_final['Call received from Registered Number (Y/N)'] == 'Y'])}")
    
    # Additional debugging info
    print(f"\nDebug Info:")
    print(f"- Order statuses in order file: {order_df['Status'].value_counts().to_dict()}")
    print(f"- Order statuses in final report: {df_final['Order Executed (Y/N)'].value_counts().to_dict()}")
    print(f"- Discussion status in final report: {df_final['Call Records matched with Order File (Y/N)'].value_counts().to_dict()}")
    
    return output_path

if __name__ == "__main__":
    import sys
    
    # Get date from command line argument or use default
    if len(sys.argv) > 1:
        date_str = sys.argv[1]
    else:
        # Default date (August 7th, 2025)
        date_str = "07082025"
    
    result = add_required_columns_for_date(date_str)
    if result:
        print(f"Successfully processed {date_str}")
    else:
        print(f"Failed to process {date_str}")
        sys.exit(1) 