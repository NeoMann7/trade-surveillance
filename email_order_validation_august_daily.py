#!/usr/bin/env python3
"""
Email-Order Validation for Daily Trade Surveillance.
Matches email trade instructions to KL orders for a given date.

Usage:
    python email_order_validation_august_daily.py 13082025
"""

import sys
import os
import json
import pandas as pd
from datetime import datetime, timedelta
import re
from openai import OpenAI
from dotenv import load_dotenv
import tempfile

# Load environment variables
load_dotenv()

# Initialize OpenAI client
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

# S3 support
USE_S3 = os.getenv('USE_S3', 'false').lower() == 'true'
S3_BUCKET_NAME = os.getenv('S3_BUCKET_NAME')
S3_BASE_PREFIX = os.getenv('S3_BASE_PREFIX', 'trade_surveillance')
SURVEILLANCE_BASE_PATH = os.getenv('SURVEILLANCE_BASE_PATH', '/app/data')

S3_AVAILABLE = False
if USE_S3 and S3_BUCKET_NAME:
    try:
        from s3_utils import (
            read_excel_from_s3, read_text_from_s3, read_csv_from_s3,
            list_s3_objects, upload_file_to_s3, get_s3_key, s3_file_exists,
            download_file_from_s3
        )
        S3_AVAILABLE = True
    except ImportError:
        try:
            from dashboard.backend.s3_utils import (
                read_excel_from_s3, read_text_from_s3, read_csv_from_s3,
                list_s3_objects, upload_file_to_s3, get_s3_key, s3_file_exists,
                download_file_from_s3
            )
            S3_AVAILABLE = True
        except ImportError:
            print("⚠️ S3 utilities not available, falling back to local filesystem")

def load_email_surveillance_results(date_str):
    """Load email surveillance results for the specific date."""
    try:
        # STANDARDIZED FORMAT: Email surveillance files are in DDMMYYYY format
        # e.g., 03102025 -> email_surveillance_{date_str}.json
        
        # Convert DDMMYYYY to YYYY-MM-DD format for S3 path
        date_obj = datetime.strptime(date_str, '%d%m%Y')
        year = date_obj.strftime('%Y')
        month_names = {
            1: "January", 2: "February", 3: "March", 4: "April",
            5: "May", 6: "June", 7: "July", 8: "August",
            9: "September", 10: "October", 11: "November", 12: "December"
        }
        month_name = month_names[date_obj.month]
        
        # CRITICAL FIX: Check local file FIRST (analyzed results from Step 2)
        # Step 2 saves analyzed results locally, and we should use those instead of S3
        # S3 should keep email_analyses (raw) for re-analysis, not all_results
        email_file = os.path.join(SURVEILLANCE_BASE_PATH, f'email_surveillance_{date_str}.json')
        if os.path.exists(email_file):
            print(f"📧 Loading email surveillance results from LOCAL file: {email_file}")
            try:
                with open(email_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # Extract trade instructions from the data structure
                # Handle both email_analyses (raw) and all_results (analyzed) formats
                if 'all_results' in data:
                    all_results = data.get('all_results', [])
                    print(f"📧 DEBUG: Found all_results with {len(all_results)} emails")
                elif 'email_analyses' in data:
                    # If we have raw emails, they need to be analyzed first
                    print(f"⚠️  Warning: Local file contains email_analyses (raw emails) but no all_results (analyzed)")
                    print(f"⚠️  Email-order validation requires analyzed emails. Please run Step 2 (AI analysis) first.")
                    all_results = []
                else:
                    print(f"⚠️  DEBUG: File structure - keys: {list(data.keys())}")
                    all_results = []
                
                # DEBUG: Check email structure and classification
                print(f"📧 DEBUG: Analyzing {len(all_results)} emails for trade instructions...")
                trade_instructions = []
                for idx, email in enumerate(all_results):
                    ai_analysis = email.get('ai_analysis', {})
                    if not ai_analysis:
                        print(f"   Email {idx+1}: No ai_analysis found")
                        continue
                    intent = ai_analysis.get('ai_email_intent', 'NOT_SET')
                    if intent == 'trade_instruction':
                        trade_instructions.append(email)
                        print(f"   Email {idx+1}: ✅ trade_instruction - {email.get('subject', '')[:60]}")
                    else:
                        print(f"   Email {idx+1}: {intent} - {email.get('subject', '')[:60]}")
                
                print(f"📧 Loaded {len(trade_instructions)} trade instructions from email surveillance")
                if len(trade_instructions) == 0 and len(all_results) > 0:
                    print(f"⚠️  WARNING: {len(all_results)} emails analyzed but 0 classified as trade_instruction!")
                    print(f"⚠️  This may indicate an issue with AI classification.")
                return data
            except Exception as e:
                print(f"❌ Error loading email file from local: {e}")
                import traceback
                traceback.print_exc()
                # Fall through to try S3
        
        # Try S3 as fallback - check Daily_Reports first (analyzed results), then Email_Data (raw)
        if USE_S3 and S3_AVAILABLE:
            # CRITICAL FIX: Check Daily_Reports first for analyzed results (all_results)
            reports_s3_key = f"{S3_BASE_PREFIX}/{month_name}/Daily_Reports/{date_str}/email_surveillance_{date_str}.json"
            if s3_file_exists(reports_s3_key):
                print(f"📧 Local file not found, trying S3 Daily_Reports: {reports_s3_key}")
                try:
                    temp_file = download_file_from_s3(reports_s3_key)
                    with open(temp_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    os.unlink(temp_file)
                    
                    if 'all_results' in data:
                        all_results = data.get('all_results', [])
                        print(f"✅ Found analyzed results (all_results) in S3 Daily_Reports: {len(all_results)} emails")
                        trade_instructions = [e for e in all_results if e.get('ai_analysis', {}).get('ai_email_intent') == 'trade_instruction']
                        print(f"📧 Loaded {len(trade_instructions)} trade instructions from S3")
                        return data
                    else:
                        print(f"⚠️  S3 Daily_Reports file exists but doesn't contain all_results")
                except Exception as e:
                    print(f"❌ Error loading from S3 Daily_Reports: {e}")
                    # Fall through to try Email_Data
            
            # Fallback: Try Email_Data (should contain email_analyses, not all_results)
            s3_key = f"{S3_BASE_PREFIX}/Email_Data/{year}/{month_name}/email_surveillance_{date_str}.json"
            if s3_file_exists(s3_key):
                print(f"📧 Daily_Reports not found, trying S3 Email_Data: {s3_key}")
                print(f"⚠️  Note: Email_Data contains email_analyses (raw), but we need all_results (analyzed)")
                print(f"⚠️  This means Step 2 (AI analysis) may not have completed successfully")
                try:
                    temp_file = download_file_from_s3(s3_key)
                    with open(temp_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    os.unlink(temp_file)
                    
                    # Extract trade instructions from the data structure
                    # Handle both email_analyses (raw) and all_results (analyzed) formats
                    if 'all_results' in data:
                        all_results = data.get('all_results', [])
                        print(f"📧 DEBUG: Found all_results with {len(all_results)} emails")
                    elif 'email_analyses' in data:
                        # If we have raw emails, they need to be analyzed first
                        print(f"⚠️  Warning: File contains email_analyses (raw emails) but no all_results (analyzed)")
                        print(f"⚠️  Email-order validation requires analyzed emails. Please run Step 2 (AI analysis) first.")
                        all_results = []
                    else:
                        print(f"⚠️  DEBUG: File structure - keys: {list(data.keys())}")
                        all_results = []
                    
                    # DEBUG: Check email structure and classification
                    print(f"📧 DEBUG: Analyzing {len(all_results)} emails for trade instructions...")
                    trade_instructions = []
                    for idx, email in enumerate(all_results):
                        ai_analysis = email.get('ai_analysis', {})
                        if not ai_analysis:
                            print(f"   Email {idx+1}: No ai_analysis found")
                            continue
                        intent = ai_analysis.get('ai_email_intent', 'NOT_SET')
                        if intent == 'trade_instruction':
                            trade_instructions.append(email)
                            print(f"   Email {idx+1}: ✅ trade_instruction - {email.get('subject', '')[:60]}")
                        else:
                            print(f"   Email {idx+1}: {intent} - {email.get('subject', '')[:60]}")
                    
                    print(f"📧 Loaded {len(trade_instructions)} trade instructions from email surveillance")
                    if len(trade_instructions) == 0 and len(all_results) > 0:
                        print(f"⚠️  WARNING: {len(all_results)} emails analyzed but 0 classified as trade_instruction!")
                        print(f"⚠️  This may indicate an issue with AI classification.")
                    return data
                except Exception as e:
                    print(f"❌ Error loading email file from S3: {e}")
                    import traceback
                    traceback.print_exc()
                    return None
            else:
                print(f"⚠️ Email surveillance file not found in S3: {s3_key}")
        
        # Try local filesystem (fallback)
        if not os.path.exists(email_file):
            # Try current directory
            email_file = f"email_surveillance_{date_str}.json"
        if not os.path.exists(email_file):
            # Try current directory
            email_file = f"email_surveillance_{date_str}.json"
        
        if os.path.exists(email_file):
            print(f"📧 Loading email surveillance results from: {email_file}")
            with open(email_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Extract trade instructions from the data structure
            # Handle both email_analyses (raw) and all_results (analyzed) formats
            if 'all_results' in data:
                all_results = data.get('all_results', [])
                print(f"📧 DEBUG: Found all_results with {len(all_results)} emails")
            elif 'email_analyses' in data:
                # If we have raw emails, they need to be analyzed first
                print(f"⚠️  Warning: File contains email_analyses (raw emails) but no all_results (analyzed)")
                print(f"⚠️  Email-order validation requires analyzed emails. Please run AI analysis first.")
                all_results = []
            else:
                print(f"⚠️  DEBUG: File structure - keys: {list(data.keys())}")
                all_results = []
            
            # DEBUG: Check email structure and classification
            print(f"📧 DEBUG: Analyzing {len(all_results)} emails for trade instructions...")
            trade_instructions = []
            for idx, email in enumerate(all_results):
                ai_analysis = email.get('ai_analysis', {})
                if not ai_analysis:
                    print(f"   Email {idx+1}: No ai_analysis found")
                    continue
                intent = ai_analysis.get('ai_email_intent', 'NOT_SET')
                if intent == 'trade_instruction':
                    trade_instructions.append(email)
                    print(f"   Email {idx+1}: ✅ trade_instruction - {email.get('subject', '')[:60]}")
                else:
                    print(f"   Email {idx+1}: {intent} - {email.get('subject', '')[:60]}")
            
            print(f"📧 Loaded {len(trade_instructions)} trade instructions from email surveillance")
            if len(trade_instructions) == 0 and len(all_results) > 0:
                print(f"⚠️  WARNING: {len(all_results)} emails analyzed but 0 classified as trade_instruction!")
                print(f"⚠️  This may indicate an issue with AI classification.")
            return data
        else:
            print(f"❌ Email surveillance file not found: {email_file}")
            if USE_S3 and S3_AVAILABLE:
                print(f"💡 Tried S3: {S3_BASE_PREFIX}/Email_Data/{year}/{month_name}/email_surveillance_{date_str}.json")
            return None
            
    except Exception as e:
        print(f"❌ Error loading email surveillance results: {e}")
        import traceback
        traceback.print_exc()
        return None

def filter_emails_by_date(email_data, target_date):
    """Filter emails by date patterns in subject."""
    filtered_emails = []
    
    # Extract trade instructions from the email data
    emails = email_data.get('trade_instructions', {}).get('emails', [])
    
    # Convert target_date to different formats
    day = target_date[:2]
    month = target_date[2:4]
    year = target_date[4:]
    
    date_patterns = [
        f"{day} Aug {year}",  # 01 Aug 25
        f"{day}/08/{year}",   # 01/08/25
        f"{day}-08-{year}",   # 01-08-25
        f"{day}{month}{year}", # 010825
        f"{day}{month}25",    # 010825
        f"{day}{month}{year[2:]}"  # 010825
    ]
    
    for email in emails:
        subject = email.get('subject', '').lower()
        
        # Check if any date pattern is in the subject
        for pattern in date_patterns:
            if pattern.lower() in subject:
                filtered_emails.append(email)
                break
    
    return filtered_emails

def load_kl_orders(date_str):
    """Load KL orders from the order file for the given date."""
    try:
        # Determine month from date string (DDMMYYYY format)
        month = int(date_str[2:4])
        month_names = {
            1: "January", 2: "February", 3: "March", 4: "April",
            5: "May", 6: "June", 7: "July", 8: "August",
            9: "September", 10: "October", 11: "November", 12: "December"
        }
        
        if month not in month_names:
            print(f"❌ Invalid month: {month}")
            return None
        
        month_name = month_names[month]
        
        # Try S3 first if enabled
        if USE_S3 and S3_AVAILABLE:
            # Try multiple possible S3 paths
            order_file_patterns = [
                f"{S3_BASE_PREFIX}/{month_name}/Order Files/OrderBook-Closed-{date_str}.csv",
                f"{S3_BASE_PREFIX}/{month_name}/Daily_Reports/{date_str}/OrderBook-Closed-{date_str}.csv",
                f"{S3_BASE_PREFIX}/{month_name}/Order Files/OrderBook_Closed-{date_str}.csv",
            ]
            
            order_file_s3_key = None
            for pattern_key in order_file_patterns:
                if s3_file_exists(pattern_key):
                    order_file_s3_key = pattern_key
                    print(f"📊 Found order file in S3: {order_file_s3_key}")
                    break
            
            if order_file_s3_key:
                try:
                    df = read_csv_from_s3(order_file_s3_key)
                    print(f"📊 Loaded {len(df)} orders from S3")
                    
                    # Filter for KL orders (User column starts with 'KL')
                    kl_orders = df[df['User'].str.startswith('KL', na=False)]
                    print(f"📊 Filtered to {len(kl_orders)} KL orders")
                    return kl_orders
                except Exception as e:
                    print(f"❌ Error loading order file from S3: {e}")
                    import traceback
                    traceback.print_exc()
                    return None
            else:
                print(f"⚠️ Order file not found in S3 for any pattern: {order_file_patterns}")
        
        # Try local filesystem
        order_file = f"{month_name}/Order Files/OrderBook-Closed-{date_str}.csv"
        print(f"📊 Loading KL orders from: {order_file}")
        
        if not os.path.exists(order_file):
            # Try alternative path
            alt_order_file = f"{month_name}/Daily_Reports/{date_str}/OrderBook-Closed-{date_str}.csv"
            if os.path.exists(alt_order_file):
                order_file = alt_order_file
            else:
                print(f"❌ Order file not found: {order_file} or {alt_order_file}")
                return None
        
        df = pd.read_csv(order_file)
        
        # Filter for KL orders (User column starts with 'KL')
        kl_orders = df[df['User'].str.startswith('KL', na=False)]
        
        print(f"📊 Loaded {len(kl_orders)} KL orders from {order_file}")
        return kl_orders
    except Exception as e:
        print(f"❌ Error loading KL orders: {e}")
        import traceback
        traceback.print_exc()
        return None

def extract_client_code_from_email(email):
    """Extract client code from email analysis - handle multiple instructions."""
    ai_analysis = email.get('ai_analysis', {})
    order_details_list = ai_analysis.get('ai_order_details', [])
    
    # Handle both single instruction (dict) and multiple instructions (list)
    if isinstance(order_details_list, dict):
        order_details_list = [order_details_list]
    elif not isinstance(order_details_list, list):
        order_details_list = []
    
    # Try to get client code from any instruction
    for order_details in order_details_list:
        client_code = order_details.get('client_code')
        if client_code:
            return client_code.upper()
    
    # Try to extract from subject if not in order details
    subject = email.get('subject', '')
    
    # Look for NEOWM patterns
    neowm_match = re.search(r'NEOWM\d+', subject, re.IGNORECASE)
    if neowm_match:
        return neowm_match.group().upper()
    
    # Look for NEO patterns
    neo_match = re.search(r'NEO\d+', subject, re.IGNORECASE)
    if neo_match:
        return neo_match.group().upper()
    
    return None

def parse_email_time(email):
    """Parse email time from order details."""
    ai_analysis = email.get('ai_analysis', {})
    order_details = ai_analysis.get('ai_order_details', {})
    
    order_time = order_details.get('order_time')
    if not order_time:
        return None
    
    try:
        # Try different time formats
        time_formats = [
            "%d%m%y %H:%M",  # 010825 09:15
            "%d/%m/%y %H:%M",  # 01/08/25 09:15
            "%d-%m-%y %H:%M",  # 01-08-25 09:15
        ]
        
        for fmt in time_formats:
            try:
                return datetime.strptime(order_time, fmt)
            except ValueError:
                continue
        
        return None
    except Exception:
        return None

def group_emails_by_instruction(emails):
    """
    Group emails by client_code + symbol
    Handle multiple instructions per email
    Deduplicate identical instructions from different emails
    """
    groups = {}
    
    for email in emails:
        client_code = extract_client_code_from_email(email)
        if not client_code:
            continue
            
        ai_analysis = email.get('ai_analysis', {})
        order_details_list = ai_analysis.get('ai_order_details', [])
        
        # Handle both single instruction (dict) and multiple instructions (list)
        if isinstance(order_details_list, dict):
            order_details_list = [order_details_list]
        elif not isinstance(order_details_list, list):
            order_details_list = []
        
        # Process each instruction in the email
        for order_details in order_details_list:
            symbol = order_details.get('symbol')
            
            if not symbol:
                continue
            
            # Create group key: client_code + symbol
            # This groups identical instructions from different emails together
            key = f"{client_code}_{symbol}"
            
            if key not in groups:
                groups[key] = []
            groups[key].append(email)
    
    print(f"📧 Grouped {len(emails)} emails into {len(groups)} instruction groups")
    return groups

def extract_final_instruction(email_group, group_key=None):
    """
    From a group of emails (same client + symbol), 
    extract ALL instructions for that client+symbol combination
    Handle multiple instructions per email
    """
    if not email_group:
        return None
    
    # If we have a group_key, extract the target symbol from it
    target_symbol = None
    if group_key:
        # Group key format: "client_code_symbol"
        parts = group_key.split('_', 1)
        if len(parts) > 1:
            target_symbol = parts[1]
    
    # Collect instructions from ALL emails in the group
    all_matching_instructions = []
    client_code = None
    
    for email in email_group:
        ai_analysis = email.get('ai_analysis', {})
        order_details_list = ai_analysis.get('ai_order_details', [])
        
        # Handle both single instruction (dict) and multiple instructions (list)
        if isinstance(order_details_list, dict):
            order_details_list = [order_details_list]
        elif not isinstance(order_details_list, list):
            order_details_list = []
        
        # Filter instructions to only include those matching the target symbol
        for order_details in order_details_list:
            if target_symbol and order_details.get('symbol') == target_symbol:
                all_matching_instructions.append(order_details)
            elif not target_symbol:
                # If no target symbol, include all instructions
                all_matching_instructions.append(order_details)
        
        # Get client code from any email (they should all be the same)
        if not client_code:
            client_code = extract_client_code_from_email(email)
    
    if not all_matching_instructions:
        return None
    
    # Deduplicate identical instructions
    unique_instructions = []
    seen_signatures = set()
    
    for instruction in all_matching_instructions:
        # Create a signature for deduplication
        signature = (
            instruction.get('symbol', ''),
            instruction.get('quantity', ''),
            instruction.get('price', ''),
            instruction.get('buy_sell', ''),
            instruction.get('order_time', '')
        )
        
        if signature not in seen_signatures:
            seen_signatures.add(signature)
            unique_instructions.append(instruction)
    
    if not unique_instructions:
        return None
    
    # Use the email with highest confidence as the source email for metadata
    final_email = max(email_group, key=lambda x: int(x.get('ai_analysis', {}).get('ai_confidence_score', 0)))
    
    # Return unique instructions from ALL emails
    return {
        'group_id': f"{client_code}_{target_symbol}",
        'client_code': client_code,
        'symbol': target_symbol,
        'instructions': unique_instructions,  # Return unique instructions from ALL emails
        'source_email': final_email,
        'all_emails_in_group': email_group
    }

def calculate_match_score(email_instruction, kl_order, email_time):
    """Calculate match score between email instruction and KL order."""
    score = 0
    discrepancies = []
    
    # 1. Client Code (Mandatory - 100 points)
    if email_instruction['client_code'] == kl_order['ClientID']:
        score += 100
    else:
        discrepancies.append(f"Client Code Mismatch - Email: {email_instruction['client_code']}, Order: {kl_order['ClientID']}")
        return 0, discrepancies  # No match possible
    
    # 2. Buy/Sell Direction (Mandatory - 15 points)
    if email_instruction['buy_sell'] == kl_order['BuySell']:
        score += 15
    else:
        discrepancies.append(f"Buy/Sell Mismatch - Email: {email_instruction['buy_sell']}, Order: {kl_order['BuySell']}")
        return 0, discrepancies  # No match possible
    
    # 3. Symbol Match (20 points)
    if email_instruction['symbol'] and kl_order['Symbol']:
        if email_instruction['symbol'].upper() in kl_order['Symbol'].upper():
            score += 20
        else:
            discrepancies.append(f"Symbol Mismatch - Email: {email_instruction['symbol']}, Order: {kl_order['Symbol']}")
    
    # 4. Quantity Match (25 points) - EXACT MATCH ONLY
    # Note: email_instruction now contains an 'instructions' array, we need to check each instruction
    for individual_instruction in email_instruction.get('instructions', []):
        if individual_instruction.get('quantity') and kl_order['Qty']:
            if individual_instruction['quantity'] == kl_order['Qty']:
                score += 25
            else:
                discrepancies.append(f"Quantity Mismatch - Email: {individual_instruction['quantity']}, Order: {kl_order['Qty']}")
            break  # Only check the first instruction for now
    
    # 5. Price Match (20 points) - Handle CMP and exact matches
    if email_instruction['price'] and kl_order['Price']:
        email_price = str(email_instruction['price']).upper()
        order_price = str(kl_order['Price'])
        
        # Handle CMP (Current Market Price) - match with any actual price but flag as discrepancy
        if email_price in ['CMP', 'CURRENT MARKET PRICE', 'MARKET PRICE']:
            score += 15  # Partial points for CMP match
            discrepancies.append(f"Price: CMP vs Actual Price {order_price}")
        elif email_price == order_price:
            score += 20  # Full points for exact match
        else:
            discrepancies.append(f"Price Mismatch - Email: {email_instruction['price']}, Order: {kl_order['Price']}")
    
    return score, discrepancies

def get_match_status(score, discrepancies):
    """Get match status based on score and discrepancies."""
    if score >= 180:
        return "Perfect Match"
    elif score >= 160:
        return "High Confidence Match"
    elif score >= 140:
        return "Time-Based Match"
    elif score >= 120:
        return "Basic Match"
    elif score >= 115:
        # Check if it's a CMP price discrepancy
        if discrepancies and any("CMP vs Actual Price" in d for d in discrepancies):
            return "CMP Price Match"
        elif discrepancies:
            return f"Partial Match - {discrepancies[0]}"
        else:
            return "Partial Match"
    else:
        return "No Match"

def convert_score_to_percentage(score):
    """Convert numerical score (0-180) to percentage (0-100%)."""
    if score == 0:
        return 0
    return round((score / 180) * 100, 1)

def match_instruction_to_orders(email_instruction, kl_orders):
    """
    Match a single email instruction to available orders
    Returns the best match with comprehensive flagging
    """
    # Filter orders by client and symbol
    matching_orders = kl_orders[
        (kl_orders['ClientID'] == email_instruction['client_code']) & 
        (kl_orders['Symbol'] == email_instruction['symbol'])
    ].copy()
    
    if len(matching_orders) == 0:
        return {
            'email_instruction': email_instruction,
            'matched_orders': [],
            'match_type': 'NO_MATCH',
                'confidence_score': 0,
            'discrepancies': [f'No orders found for client {email_instruction["client_code"]} and symbol {email_instruction["symbol"]}'],
            'review_flags': []
        }
    
    # Calculate total quantity from all matching orders
    total_order_qty = matching_orders['Qty'].sum()
    # Get quantity from first instruction in the group
    email_qty = None
    for individual_instruction in email_instruction.get('instructions', []):
        if individual_instruction.get('quantity'):
            email_qty = individual_instruction['quantity']
            break
    
    # Find the best matching order(s)
    best_match = None
    best_score = 0
    best_discrepancies = []
    
    for _, order in matching_orders.iterrows():
        score, discrepancies = calculate_match_score(email_instruction, order, None)
        if score > best_score:
            best_score = score
            best_match = order
            best_discrepancies = discrepancies
            
    # Generate review flags
    review_flags = []
    
    # Flag 1: Quantity Mismatch
    if total_order_qty != email_qty:
        review_flags.append({
            'type': 'QUANTITY_MISMATCH',
            'email_quantity': email_qty,
            'total_order_quantity': total_order_qty,
            'difference': int(total_order_qty) - int(email_qty) if email_qty and total_order_qty else 0
        })
    
    # Flag 2: Price Mismatch
    price_mismatches = []
    for _, order in matching_orders.iterrows():
        if order['Price'] != email_instruction['price']:
            # Handle non-numeric price values
            try:
                order_price_float = float(order['Price']) if order['Price'] else 0
                email_price_float = float(email_instruction['price']) if email_instruction['price'] else 0
                price_difference = order_price_float - email_price_float
            except (ValueError, TypeError):
                price_difference = 0
            
            price_mismatches.append({
                'order_id': order['NorenOrderID'],
                'email_price': email_instruction['price'],
                'order_price': order['Price'],
                'difference': price_difference
            })
    
    if price_mismatches:
        review_flags.append({
            'type': 'PRICE_MISMATCH',
            'details': price_mismatches
        })
    
    # Flag 3: Multiple Orders
    if len(matching_orders) > 1:
        review_flags.append({
            'type': 'MULTIPLE_ORDERS',
            'order_count': len(matching_orders),
            'orders': matching_orders['NorenOrderID'].tolist()
        })
    
    # Convert quantities to integers for comparison
    try:
        total_order_qty_int = int(total_order_qty) if total_order_qty else 0
        email_qty_int = int(email_qty) if email_qty else 0
    except (ValueError, TypeError):
        total_order_qty_int = 0
        email_qty_int = 0
    
    # Determine match type
    if len(matching_orders) == 1 and total_order_qty_int == email_qty_int and best_score >= 180:
        match_type = 'PERFECT_MATCH'
    elif len(matching_orders) > 1 and total_order_qty_int == email_qty_int:
        match_type = 'SPLIT_EXECUTION'
    elif total_order_qty_int < email_qty_int:
        match_type = 'PARTIAL_MATCH'
    elif total_order_qty_int > email_qty_int:
        match_type = 'OVER_MATCH'
    else:
        match_type = 'BASIC_MATCH'
    
    return {
        'email_instruction': email_instruction,
        'matched_orders': matching_orders.to_dict('records'),
        'match_type': match_type,
        'confidence_score': convert_score_to_percentage(best_score),
        'discrepancies': best_discrepancies,
        'review_flags': review_flags
    }

def match_email_group_to_orders_with_ai(email_group, kl_orders, group_key=None):
    """
    Use AI to match email group to orders with intelligent symbol matching
    """
    print(f"🔍 DEBUG: match_email_group_to_orders_with_ai called with group_key: {group_key}")
    # Get the final instruction from the email group
    final_instruction = extract_final_instruction(email_group, group_key)
    print(f"🔍 DEBUG: Final instruction extracted: {final_instruction}")
    
    # Filter orders by client code first - with normalization
    client_code = final_instruction['client_code']
    
    # Try exact match first
    client_orders = kl_orders[kl_orders['ClientID'] == client_code].copy()
    
    # If no exact match, try normalized client code patterns
    if len(client_orders) == 0:
        # Handle EOWM -> NEOWM normalization
        if client_code.startswith('EOWM') and not client_code.startswith('NEOWM'):
            normalized_code = f'NEOWM{client_code[4:]}'  # EOWM00542 -> NEOWM00542
            client_orders = kl_orders[kl_orders['ClientID'] == normalized_code].copy()
            if len(client_orders) > 0:
                print(f"🔍 DEBUG: Found orders with normalized client code: {normalized_code}")
        
        # Handle WM -> NEOWM normalization
        elif client_code.startswith('WM') and not client_code.startswith('NEOWM'):
            normalized_code = f'NEO{client_code}'
            client_orders = kl_orders[kl_orders['ClientID'] == normalized_code].copy()
            if len(client_orders) > 0:
                print(f"🔍 DEBUG: Found orders with normalized client code: {normalized_code}")
        
        # Handle other NEO patterns
        elif not client_code.startswith('NEO') and client_code.isdigit():
            # Try NEO + client_code
            normalized_code = f'NEO{client_code}'
            client_orders = kl_orders[kl_orders['ClientID'] == normalized_code].copy()
            if len(client_orders) > 0:
                print(f"🔍 DEBUG: Found orders with normalized client code: {normalized_code}")
    
    print(f"🔍 DEBUG: Client orders found for {client_code}: {len(client_orders)}")
    
    if len(client_orders) == 0:
        print(f"🔍 DEBUG: ❌ No orders found for client {client_code}")
        return {
            'email_instruction': final_instruction,
            'matched_orders': [],
            'match_type': 'NO_MATCH',
            'confidence_score': 0,
            'discrepancies': [f'No orders found for client {client_code}'],
            'review_flags': []
        }
    
    # Prepare ALL email instructions for AI
    email_instructions = []
    for instruction in final_instruction['instructions']:
        email_instructions.append({
            'client_code': final_instruction['client_code'],
            'symbol': instruction.get('symbol'),
            'quantity': instruction.get('quantity'),
            'price': instruction.get('price'),
            'buy_sell': instruction.get('buy_sell'),
            'order_time': instruction.get('order_time', ''),
            'subject': email_group[0]['subject']
        })
    
    # Prepare available orders for AI
    available_orders = []
    for _, order in client_orders.iterrows():
        available_orders.append({
            'order_id': str(order['NorenOrderID']),
            'symbol': order['Symbol'],
            'quantity': int(order['Qty']),
            'price': float(order['Price']) if order['Price'] else 0,
            'buy_sell': order['BuySell'],
            'status': order.get('Status', '')
        })
    
    # Create AI prompt for matching
    prompt = f"""
    You are a trade surveillance expert. Match the email trade instructions to the available orders.

    **EMAIL TRADE INSTRUCTIONS:**
    {json.dumps(email_instructions, indent=2)}

    **AVAILABLE ORDERS FOR CLIENT {final_instruction['client_code']}:**
    {json.dumps(available_orders, indent=2)}

    **CRITICAL MATCHING RULES:**

    1. **SYMBOL MATCHING** - Be intelligent about variations:
       - "blue jet healthcare" = "BLUEJET"
       - "Energy INVIT" = "ENERGYINF"
       - "Manappuram Finance Limited" = "MANAPPURAM"

    2. **VALUE-BASED QUANTITY MATCHING**:
       - If email says "Worth INR X" → Calculate total value of ALL matching orders
       - Example: Email "Worth INR 5,00,000" should match orders totaling ~₹5,00,000
       - Formula: Sum(Quantity × Price) for all matching orders

    3. **SPLIT EXECUTION LOGIC**:
       - If multiple orders together fulfill the instruction → Match ALL of them
       - Don't just pick the largest order - include ALL relevant orders
       - Example: 3 orders (10+20+710=740 shares) = ₹493,580 ≈ ₹5,00,000

        4. **ORDER STATUS HANDLING**:
           - Match orders regardless of status (Complete, Active, Cancelled, Rejected)
           - All orders provide valuable surveillance data
           - Flag status discrepancies for review if needed

        5. **PRICE DISCREPANCY HANDLING**:
           - "Market price"/"CMP" → Match with actual price but flag as discrepancy
           - Flag: "Price: Market price vs Actual Price ₹667"
           
        6. **STATUS DISCREPANCY HANDLING**:
           - If order status is not "Complete", flag for review
           - Flag: "Order Status: Cancelled/Rejected (order was placed but not completed)"

    7. **BUY/SELL DIRECTION**: Must match exactly

    **IMPORTANT:** You will receive multiple instructions for the same client+symbol. Match each instruction to the appropriate orders. If multiple instructions can be fulfilled by the same orders, that's fine - just ensure each instruction gets matched to the most appropriate orders.

    **RETURN JSON:**
    {{
        "matched_order_ids": ["list of order IDs that match ANY of the instructions"],
        "confidence_score": 0-100,
        "reasoning": "explanation of the matches for all instructions",
        "match_type": "EXACT_MATCH|SPLIT_EXECUTION|PARTIAL_MATCH|NO_MATCH",
        "discrepancies": ["list of any discrepancies found"],
        "review_required": true/false
    }}
    """
    
    try:
        print(f"🔍 DEBUG: Sending prompt to AI...")
        print(f"🔍 DEBUG: Available orders for AI: {len(available_orders)}")
        for order in available_orders:
            print(f"🔍 DEBUG:   Order: {order['order_id']} - {order['symbol']} - {order['quantity']} - {order['price']}")
        
        response = client.chat.completions.create(
            model="gpt-4.1",
            messages=[
                {"role": "system", "content": "You are a trade surveillance expert. Match trade instructions to orders accurately."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )
        
        ai_response = response.choices[0].message.content.strip()
        print(f"🔍 DEBUG: AI Response: {ai_response[:200]}...")
        
        # Parse AI response - handle markdown code blocks and extra content
        if ai_response.startswith('```json'):
            ai_response = ai_response[7:]
        if ai_response.endswith('```'):
            ai_response = ai_response[:-3]
        
        # Find the JSON part (between first { and last })
        start_idx = ai_response.find('{')
        end_idx = ai_response.rfind('}')
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            json_part = ai_response[start_idx:end_idx+1]
            ai_result = json.loads(json_part.strip())
        else:
            ai_result = json.loads(ai_response.strip())
        print(f"🔍 DEBUG: Parsed AI result: {ai_result}")
        
        # Get matched orders
        matched_orders = []
        for order_id in ai_result.get('matched_order_ids', []):
            print(f"🔍 DEBUG: Looking for order ID: {order_id}")
            matched_order = client_orders[client_orders['NorenOrderID'] == float(order_id)]
            if not matched_order.empty:
                matched_orders.append(matched_order.iloc[0].to_dict())
                print(f"🔍 DEBUG: ✅ Found matched order: {order_id}")
            else:
                print(f"🔍 DEBUG: ❌ Order not found: {order_id}")
        
        # Generate review flags
        review_flags = []
        if ai_result.get('review_required', False):
            review_flags.append({
                'type': 'AI_REVIEW_REQUIRED',
                'reason': ai_result.get('reasoning', 'AI flagged for review')
            })
        
        return {
            'email_instruction': final_instruction,
            'matched_orders': matched_orders,
            'match_type': ai_result.get('match_type', 'NO_MATCH'),
            'confidence_score': ai_result.get('confidence_score', 0),
            'discrepancies': ai_result.get('discrepancies', []),
            'review_flags': review_flags
        }
        
    except Exception as e:
        print(f"❌ AI matching failed: {e}")
        # Fallback to exact matching
        return match_instruction_to_orders(final_instruction, kl_orders)

def assign_emails_to_orders(emails, kl_orders):
    """
    NEW LOGIC: Group emails by instruction and match to orders
    Ensures one-to-one mapping between email instructions and order groups
    """
    print(f"🔄 Starting new email-to-order mapping logic...")
    print(f"🔍 DEBUG: Input emails count: {len(emails)}")
    print(f"🔍 DEBUG: Input KL orders count: {len(kl_orders)}")
    
    # Step 1: Group emails by instruction
    email_groups = group_emails_by_instruction(emails)
    print(f"🔍 DEBUG: After grouping - email_groups count: {len(email_groups)}")
    for group_key, group_emails in email_groups.items():
        print(f"🔍 DEBUG: Group '{group_key}' has {len(group_emails)} emails")
    
    # Step 2: Extract final instruction from each group
    final_instructions = []
    for group_key, email_group in email_groups.items():
        print(f"🔍 DEBUG: Processing group '{group_key}' with {len(email_group)} emails")
        final_instruction = extract_final_instruction(email_group, group_key)
        if final_instruction:
            print(f"🔍 DEBUG: ✅ Extracted instruction for '{group_key}': {final_instruction.get('symbol')} - {final_instruction.get('client_code')}")
            final_instructions.append(final_instruction)
        else:
            print(f"🔍 DEBUG: ❌ Failed to extract instruction for '{group_key}'")
    
    print(f"📧 Extracted {len(final_instructions)} final instructions")
    
    # Step 3: Match each instruction to orders
    assignments = []
    used_orders = set()
    
    print(f"🔍 DEBUG: Starting matching for {len(email_groups)} groups")
    for group_key, email_group in email_groups.items():
        print(f"🔍 DEBUG: Matching group '{group_key}' with {len(email_group)} emails")
        match_result = match_email_group_to_orders_with_ai(email_group, kl_orders, group_key)
        print(f"🔍 DEBUG: Match result for '{group_key}': {match_result.get('match_type')} - {match_result.get('confidence_score')}%")
        print(f"🔍 DEBUG: Matched orders count: {len(match_result.get('matched_orders', []))}")
        if match_result.get('matched_orders'):
            for order in match_result['matched_orders']:
                print(f"🔍 DEBUG:   Matched order: {order.get('NorenOrderID')} - {order.get('symbol')}")
        
        # Check if any of the matched orders are already used
        matched_order_ids = [order['NorenOrderID'] for order in match_result['matched_orders']]
        already_used = any(order_id in used_orders for order_id in matched_order_ids)
        
        if already_used:
            # Mark as conflict
            match_result['match_type'] = 'ORDER_CONFLICT'
            match_result['discrepancies'].append('Order already assigned to another instruction')
        
        # Add matched_order_ids to the result structure
        match_result['matched_order_ids'] = matched_order_ids
        
        # Mark orders as used
        for order_id in matched_order_ids:
            used_orders.add(order_id)
        
        assignments.append(match_result)
    
    print(f"\n📊 NEW MAPPING SUMMARY:")
    print(f"   Total email groups: {len(email_groups)}")
    print(f"   Final instructions: {len(final_instructions)}")
    print(f"   Successful matches: {len([a for a in assignments if a['match_type'] != 'NO_MATCH'])}")
    print(f"   Orders used: {len(used_orders)}")
    
    return assignments

def match_emails_to_orders(emails, kl_orders):
    """Match email trade instructions to KL orders using new logic."""
    return assign_emails_to_orders(emails, kl_orders)

def generate_mapping_report(matches, date_str):
    """Generate email-order mapping report."""
    report_data = []
    
    for match in matches:
        instruction = match['email_instruction']
        orders = match['matched_orders']
        
        # Get primary order for reporting
        primary_order = orders[0] if orders else {}
        
        # Handle new structure where instruction contains multiple instructions
        if 'instructions' in instruction and instruction['instructions']:
            # Use the first instruction for reporting
            first_instruction = instruction['instructions'][0]
            report_row = {
                'Email_Group_ID': instruction['group_id'],
                'Email_Client_Code': instruction['client_code'],
                'Email_Symbol': instruction['symbol'],
                'Email_Quantity': first_instruction.get('quantity', ''),
                'Email_Price': first_instruction.get('price', ''),
                'Email_Buy_Sell': first_instruction.get('buy_sell', ''),
                'Match_Type': match['match_type'],
                'Confidence_Score': match['confidence_score'],
                'Total_Order_Quantity': sum(order['Qty'] for order in orders),
                'Order_Count': len(orders),
                'Order_IDs': ','.join([str(order['NorenOrderID']) for order in orders]),
                'Review_Required': 'YES' if match['review_flags'] else 'NO',
                'Flag_Count': len(match['review_flags']),
                'Flags': '; '.join([flag['type'] for flag in match['review_flags']]),
                'Discrepancies': '; '.join(match['discrepancies']) if match['discrepancies'] else 'None'
            }
        else:
            # Fallback to old structure
            report_row = {
                'Email_Group_ID': instruction['group_id'],
                'Email_Client_Code': instruction['client_code'],
                'Email_Symbol': instruction['symbol'],
                'Email_Quantity': instruction.get('quantity', ''),
                'Email_Price': instruction.get('price', ''),
                'Email_Buy_Sell': instruction.get('buy_sell', ''),
                'Match_Type': match['match_type'],
                'Confidence_Score': match['confidence_score'],
                'Total_Order_Quantity': sum(order['Qty'] for order in orders),
                'Order_Count': len(orders),
                'Order_IDs': ','.join([str(order['NorenOrderID']) for order in orders]),
                'Review_Required': 'YES' if match['review_flags'] else 'NO',
                'Flag_Count': len(match['review_flags']),
                'Flags': '; '.join([flag['type'] for flag in match['review_flags']]),
                'Discrepancies': '; '.join(match['discrepancies']) if match['discrepancies'] else 'None'
            }
        
        report_data.append(report_row)
    
    # Create DataFrame and save
    df = pd.DataFrame(report_data)
    
    # Determine month from date
    day = date_str[:2]
    month = date_str[2:4]
    year = date_str[4:8]
    
    # Map month number to month name
    month_names = {
        '01': 'January', '02': 'February', '03': 'March', '04': 'April',
        '05': 'May', '06': 'June', '07': 'July', '08': 'August',
        '09': 'September', '10': 'October', '11': 'November', '12': 'December'
    }
    month_name = month_names.get(month, 'Unknown')
    
    output_file = f"{month_name}/Daily_Reports/{date_str}/email_order_mapping_{date_str}.xlsx"
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    df.to_excel(output_file, index=False)
    print(f"📊 Email-order mapping report saved: {output_file}")
    
    return df

def save_mapping_data(matches, date_str):
    """Save mapping data for use by Excel generation script."""
    mapping_data = {
        'date': date_str,
        'timestamp': datetime.now().strftime('%Y%m%d_%H%M%S'),
        'total_instructions': len(matches),
        'matched_instructions': len([m for m in matches if m['match_type'] != 'NO_MATCH']),
        'unmatched_instructions': len([m for m in matches if m['match_type'] == 'NO_MATCH']),
        'matches': matches
    }
    
    # Determine month from date
    day = date_str[:2]
    month = date_str[2:4]
    year = date_str[4:8]
    
    # Map month number to month name
    month_names = {
        '01': 'January', '02': 'February', '03': 'March', '04': 'April',
        '05': 'May', '06': 'June', '07': 'July', '08': 'August',
        '09': 'September', '10': 'October', '11': 'November', '12': 'December'
    }
    month_name = month_names.get(month, 'Unknown')
    
    output_file = f"{month_name}/Daily_Reports/{date_str}/email_order_mapping_{date_str}.json"
    
    if USE_S3 and S3_AVAILABLE:
        # Save to S3
        s3_key = f"{S3_BASE_PREFIX}/{output_file}"
        # Create temporary file to upload
        import tempfile
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp_file:
            json.dump(mapping_data, tmp_file, indent=2, default=str)
            tmp_file_path = tmp_file.name
        
        try:
            upload_file_to_s3(tmp_file_path, s3_key)
            print(f"💾 Email-order mapping data saved to S3: {s3_key}")
        finally:
            # Clean up temp file
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
    else:
        # Save locally
        # Ensure directory exists
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        with open(output_file, 'w') as f:
            json.dump(mapping_data, f, indent=2, default=str)
        
        print(f"💾 Email-order mapping data saved: {output_file}")
    
    return mapping_data

def normalize_order_id(order_id):
    """Convert scientific notation to full order ID for comparison"""
    if pd.isna(order_id):
        return None
    try:
        # Convert scientific notation to full number, then to string
        normalized = str(int(float(order_id)))
        # Handle precision issues for very large numbers
        # If the normalized ID is very close to the original, use the original
        if len(normalized) > 15:  # Very large order IDs
            # Try to match by checking if they're within a small range
            original_int = int(float(order_id))
            return str(original_int)
        return normalized
    except (ValueError, TypeError):
        return None

def update_audio_surveillance_excel(matches, date_str):
    """Update the audio surveillance Excel file with email-order mapping information."""
    print("📊 Step 6: Updating audio surveillance Excel with email-order mapping...")
    
    # Find the audio surveillance Excel file - determine month from date
    day = date_str[:2]
    month = date_str[2:4]
    year = date_str[4:8]
    
    # Map month number to month name
    month_names = {
        '01': 'January', '02': 'February', '03': 'March', '04': 'April',
        '05': 'May', '06': 'June', '07': 'July', '08': 'August',
        '09': 'September', '10': 'October', '11': 'November', '12': 'December'
    }
    month_name = month_names.get(month, 'Unknown')
    
    audio_file = f"{month_name}/Daily_Reports/{date_str}/Final_Trade_Surveillance_Report_{date_str}_with_Email_and_Trade_Analysis.xlsx"
    
    if not os.path.exists(audio_file):
        print(f"⚠️  Audio surveillance file not found: {audio_file}")
        return False
    
    try:
        # Load the audio surveillance Excel file
        df = pd.read_excel(audio_file)
        print(f"📊 Loaded audio surveillance file with {len(df)} orders")
        
        # Load KL orders to create NorenOrderID to ExchOrderID mapping
        kl_orders = load_kl_orders(date_str)
        if kl_orders is not None:
            # Create mapping from NorenOrderID to ExchOrderID
            noren_to_exch_mapping = {}
            for _, row in kl_orders.iterrows():
                noren_id = row.get('NorenOrderID')
                exch_id = row.get('ExchOrderID')
                if pd.notna(noren_id) and pd.notna(exch_id):
                    noren_to_exch_mapping[noren_id] = exch_id
            print(f"📊 Created NorenOrderID to ExchOrderID mapping for {len(noren_to_exch_mapping)} orders")
        else:
            noren_to_exch_mapping = {}
            print("⚠️  Could not load KL orders for ID mapping")
        
        # Create a mapping dictionary for quick lookup with normalized order IDs
        email_order_mapping = {}
        for match in matches:
            instruction = match['email_instruction']
            orders = match['matched_orders']
            
            # Create mapping for each order in this match
            for order in orders:
                noren_order_id = order.get('NorenOrderID')
                if pd.notna(noren_order_id):  # Only map orders with valid IDs
                    # Map NorenOrderID to ExchOrderID for Excel matching
                    exch_order_id = noren_to_exch_mapping.get(noren_order_id, noren_order_id)
                    # Normalize the order ID for comparison
                    normalized_order_id = normalize_order_id(exch_order_id)
                    if normalized_order_id:
                        # Create complete email content including body and attachments
                        source_email = instruction['source_email']
                        email_content = f"Subject: {source_email.get('subject', '')}\n"
                        email_content += f"Sender: {source_email.get('sender', '')}\n"
                        email_content += f"Date: {source_email.get('date', '')}\n"
                        email_content += f"Client: {instruction.get('client_code', '')}\n"
                        email_content += f"Symbol: {instruction.get('symbol', '')}\n"
                        # Get quantity from first instruction
                        first_instruction = instruction.get('instructions', [{}])[0] if instruction.get('instructions') else {}
                        email_content += f"Qty: {first_instruction.get('quantity', '')}\n"
                        email_content += f"Price: {first_instruction.get('price', '')}\n"
                        email_content += f"Action: {first_instruction.get('buy_sell', '')}\n"
                        email_content += f"\n--- EMAIL BODY ---\n"
                        email_content += f"{source_email.get('clean_text', '')}\n"
                        
                        # Add PDF attachment content if present
                        if source_email.get('has_attachments', False) and source_email.get('attachments'):
                            email_content += f"\n--- PDF ATTACHMENTS ---\n"
                            for attachment in source_email.get('attachments', []):
                                if 'extracted_text' in attachment:
                                    email_content += f"\nPDF: {attachment.get('name', 'Unknown')}\n"
                                    email_content += f"{attachment['extracted_text']}\n"
                        
                        email_order_mapping[normalized_order_id] = {
                            'email_match': 'Matched',
                            'confidence_score': f"{match['confidence_score']}%",
                            'discrepancies': [],
                            'review_required': 'N',
                            'email_content': email_content
                        }
                        
                        # Check for discrepancies
                        if match['review_flags']:
                            email_order_mapping[normalized_order_id]['review_required'] = 'Y'
                            
                            # Build discrepancy description
                            discrepancies = []
                            for flag in match['review_flags']:
                                if flag['type'] == 'QUANTITY_MISMATCH':
                                    # Get quantity from first instruction
                                    first_instruction = instruction.get('instructions', [{}])[0] if instruction.get('instructions') else {}
                                    discrepancies.append(f"Quantity: Email={first_instruction.get('quantity', '')}, Orders={flag['total_order_quantity']}")
                                elif flag['type'] == 'PRICE_MISMATCH':
                                    for detail in flag['details']:
                                        if detail['order_id'] == order_id:
                                            discrepancies.append(f"Price: Email={detail['email_price']}, Order={detail['order_price']}")
                                elif flag['type'] == 'MULTIPLE_ORDERS':
                                    discrepancies.append(f"Multiple orders matched ({flag['order_count']} orders)")
                            
                            email_order_mapping[normalized_order_id]['discrepancies'] = discrepancies
        
        # Update existing email-order mapping columns in the DataFrame
        # Use the correct column names that exist in the Excel file
        if 'Email-Order Match Status' not in df.columns:
            df['Email-Order Match Status'] = 'No Email Match'  # Default
        if 'Email Confidence Score' not in df.columns:
            df['Email Confidence Score'] = '0%'  # Default
        if 'Email Discrepancy Details' not in df.columns:
            df['Email Discrepancy Details'] = 'No email data'  # Default
        if 'Email_Content' not in df.columns:
            df['Email_Content'] = ''  # Default to empty
        if 'Email_Review_Required' not in df.columns:
            df['Email_Review_Required'] = 'N'
        
        # Update rows based on normalized Order ID matches
        matched_count = 0
        fallback_matched_count = 0
        
        for idx, row in df.iterrows():
            order_id = row.get('Order ID')  # Audio surveillance uses 'Order ID' column
            if pd.notna(order_id):
                # Try order ID matching first
                normalized_audio_order_id = normalize_order_id(order_id)
                if normalized_audio_order_id and normalized_audio_order_id in email_order_mapping:
                    mapping = email_order_mapping[normalized_audio_order_id]
                    df.at[idx, 'Email-Order Match Status'] = mapping['email_match']
                    df.at[idx, 'Email Discrepancy Details'] = '; '.join(mapping['discrepancies']) if mapping['discrepancies'] else ''
                    df.at[idx, 'Email Confidence Score'] = mapping['confidence_score']
                    
                    # Add email content for matched orders
                    if 'email_content' in mapping:
                        df.at[idx, 'Email_Content'] = mapping['email_content']
                    
                    # Review required flag if present
                    df.at[idx, 'Email_Review_Required'] = mapping.get('review_required', 'N')
                    matched_count += 1
                else:
                    # Fallback: Try matching by client code, symbol, and quantity
                    client_code = row.get('Client Code')
                    symbol = row.get('symbol')  # Audio surveillance uses lowercase column names
                    quantity = row.get('quantity')  # Audio surveillance uses lowercase column names
                    
                    if client_code and symbol and quantity:
                        # Find matching email instruction by client, symbol, and quantity
                        for match in matches:
                            instruction_group = match['email_instruction']
                            # Check each individual instruction in the group
                            for individual_instruction in instruction_group.get('instructions', []):
                                # Convert quantities to same type for comparison
                                email_qty = str(individual_instruction['quantity']) if individual_instruction.get('quantity') else ''
                                excel_qty = str(quantity) if quantity else ''
                                
                                if (individual_instruction.get('client_code') == client_code and 
                                    individual_instruction.get('symbol') == symbol and 
                                    email_qty == excel_qty):
                                
                                    # Found a match by fallback criteria
                                    source_email = instruction_group['source_email']
                                    email_content = f"Subject: {source_email.get('subject', '')}\n"
                                    email_content += f"Sender: {source_email.get('sender', '')}\n"
                                    email_content += f"Date: {source_email.get('date', '')}\n"
                                    email_content += f"Client: {individual_instruction.get('client_code', '')}\n"
                                    email_content += f"Symbol: {individual_instruction.get('symbol', '')}\n"
                                    email_content += f"Qty: {individual_instruction.get('quantity', '')}\n"
                                    email_content += f"Price: {individual_instruction.get('price', '')}\n"
                                    email_content += f"Action: {individual_instruction.get('buy_sell', '')}\n"
                                    email_content += f"\n--- EMAIL BODY ---\n"
                                    email_content += f"{source_email.get('clean_text', '')}\n"
                                    
                                    # Add PDF attachment content if present
                                    if source_email.get('has_attachments', False) and source_email.get('attachments'):
                                        email_content += f"\n--- PDF ATTACHMENTS ---\n"
                                        for attachment in source_email.get('attachments', []):
                                            if 'extracted_text' in attachment:
                                                email_content += f"\nPDF: {attachment.get('name', 'Unknown')}\n"
                                                email_content += f"{attachment['extracted_text']}\n"
                                    
                                    mapping = {
                                        'email_match': 'Matched',
                                        'confidence_score': f"{match['confidence_score']}%",
                                        'discrepancies': match.get('discrepancies', []),
                                        'review_required': 'Y' if match.get('review_flags') else 'N',
                                        'email_content': email_content
                                    }
                                    
                                    df.at[idx, 'Email-Order Match Status'] = mapping['email_match']
                                    df.at[idx, 'Email Discrepancy Details'] = '; '.join(mapping['discrepancies']) if mapping['discrepancies'] else ''
                                    df.at[idx, 'Email Confidence Score'] = mapping['confidence_score']
                                    df.at[idx, 'Email_Content'] = mapping['email_content']
                                    df.at[idx, 'Email_Review_Required'] = mapping.get('review_required', 'N')
                                    
                                    fallback_matched_count += 1
                                    break
        
        # Add highlighting for orders with no source (no audio + no email + completed status)
        df['No_Source_Found'] = 'N'  # Default to 'N'
        
        # Check for orders with no source
        for idx, row in df.iterrows():
            email_match = row.get('Email_Order_Match', 'N')
            audio_mapped = row.get('audio_mapped', 'N')
            status = row.get('status', '').upper()
            
            # Check if order has no source and is completed (not cancelled/rejected)
            # audio_mapped can be 'N', 'no', or 'NO' - check for any negative value
            if (email_match == 'N' and 
                audio_mapped.lower() in ['n', 'no'] and 
                status not in ['CANCELLED', 'REJECTED', 'CANCEL', 'REJECT']):
                df.at[idx, 'No_Source_Found'] = 'Y'
        
        # Count orders with no source
        no_source_count = df[df['No_Source_Found'] == 'Y'].shape[0]
        
        # Save the updated file with highlighting
        with pd.ExcelWriter(audio_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Trade_Surveillance')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Trade_Surveillance']
            
            # Define highlighting styles
            from openpyxl.styles import PatternFill, Font
            
            # Red background for orders with no source
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            red_font = Font(bold=True, color='CC0000')
            
            # Apply highlighting to rows with no source
            for idx, row in df.iterrows():
                if row['No_Source_Found'] == 'Y':
                    # Highlight the entire row (columns A to Z)
                    # Use the actual Excel row number (idx + 2 because Excel is 1-indexed and we have header)
                    # Since we're using index=False, the data starts from column A (index=1)
                    excel_row = idx + 2
                    for col in range(1, 28):  # A to Z
                        cell = worksheet.cell(row=excel_row, column=col)
                        cell.fill = red_fill
                        cell.font = red_font
        
        print(f"✅ Updated audio surveillance Excel file: {audio_file}")
        
        # Print summary
        # Summary counts based on September column names
        matched_orders = df[df['Email-Order Match Status'].astype(str) != 'No Email Match'].shape[0]
        review_required = df[df['Email_Review_Required'] == 'Y'].shape[0]
        print(f"📊 Summary: {matched_orders} orders matched to emails, {review_required} require review")
        print(f"🔍 Order ID normalization: {matched_count} orders matched using normalized order IDs")
        print(f"🔄 Fallback matching: {fallback_matched_count} orders matched using client/symbol/quantity")
        print(f"⚠️  Orders with no source (highlighted in red): {no_source_count}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating audio surveillance Excel: {e}")
        return False

def main():
    if len(sys.argv) != 2:
        print("Usage: python email_order_validation_august_daily.py <date>")
        print("Example: python email_order_validation_august_daily.py 13082025")
        sys.exit(1)
    
    date_str = sys.argv[1]
    
    print(f"🔄 Starting Email-Order Validation for {date_str}")
    print(f"{'='*60}")
    
    # Step 1: Load email surveillance results
    print("📧 Step 1: Loading email surveillance results...")
    all_emails = load_email_surveillance_results(date_str)
    if all_emails is None:
        sys.exit(1)
    
    # Step 2: Get trade instructions (skip date filtering for date-specific files)
    print("📅 Step 2: Getting trade instructions...")
    # Use all_results instead of trade_instructions.emails to get complete email data with all instructions
    all_results = all_emails.get('all_results', [])
    print(f"📧 DEBUG: Total emails in all_results: {len(all_results)}")
    
    # DEBUG: Show breakdown of email intents
    intent_counts = {}
    for email in all_results:
        ai_analysis = email.get('ai_analysis', {})
        intent = ai_analysis.get('ai_email_intent', 'NOT_SET') if ai_analysis else 'NO_ANALYSIS'
        intent_counts[intent] = intent_counts.get(intent, 0) + 1
    
    print(f"📧 DEBUG: Email intent breakdown: {intent_counts}")
    
    date_emails = [email for email in all_results if email.get('ai_analysis', {}).get('ai_email_intent') == 'trade_instruction']
    print(f"📧 Found {len(date_emails)} trade instructions")
    
    if len(date_emails) == 0:
        print(f"⚠️  WARNING: No trade instructions found! This will result in 0 email matches.")
        print(f"⚠️  Check if:")
        print(f"   1. Emails were properly analyzed by complete_email_surveillance_system.py")
        print(f"   2. AI classified emails as 'trade_instruction' (not 'trade_confirmation' or 'other')")
        print(f"   3. The email_surveillance file contains 'all_results' with proper structure")
    
    # Step 3: Load KL orders
    print("📊 Step 3: Loading KL orders...")
    kl_orders = load_kl_orders(date_str)
    if kl_orders is None:
        sys.exit(1)
    
    # Step 4: Match emails to orders using new logic
    print("🔗 Step 4: Matching emails to orders...")
    matches = match_emails_to_orders(date_emails, kl_orders)
    
    # Step 5: Generate reports
    print("📋 Step 5: Generating reports...")
    mapping_df = generate_mapping_report(matches, date_str)
    mapping_data = save_mapping_data(matches, date_str)
    
    # Step 6: Update audio surveillance Excel with email-order mapping
    update_audio_surveillance_excel(matches, date_str)
    
    # Summary
    print(f"\n{'='*60}")
    print("📊 EMAIL-ORDER VALIDATION SUMMARY")
    print(f"{'='*60}")
    print(f"📅 Date: {date_str}")
    print(f"📧 Total Email Groups: {mapping_data['total_instructions']}")
    print(f"✅ Matched Instructions: {mapping_data['matched_instructions']}")
    print(f"❌ Unmatched Instructions: {mapping_data['unmatched_instructions']}")
    if mapping_data['total_instructions'] > 0:
        print(f"📈 Match Rate: {(mapping_data['matched_instructions']/mapping_data['total_instructions']*100):.1f}%")
    else:
        print(f"📈 Match Rate: 0.0% (no instructions to match)")
    
    print(f"\n🎉 Email-order validation completed successfully!")
    print(f"📁 Reports saved in: August/Daily_Reports/{date_str}/")

if __name__ == "__main__":
    main() 